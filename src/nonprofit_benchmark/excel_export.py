"""Excel Export (pure, no I/O).

Renders the dashboard's filtered view — benchmark rows plus summary
statistics — into a board-ready .xlsx workbook, returned as bytes.
"""

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from nonprofit_benchmark.benchmark import BenchmarkRow, SummaryStats

PEER_HEADER = [
    "Organization",
    "City",
    "State",
    "NTEE",
    "Total revenue",
    "Executive title",
    "Executive compensation",
    "% of revenue",
    "Filing year",
    "Data source",
    "ProPublica URL",
    "Stale (>3 yrs)",
]

PEER_COLUMN_WIDTHS = {
    "Organization": 40,
    "ProPublica URL": 60,
}
DEFAULT_COLUMN_WIDTH = 22

BOLD = Font(bold=True)


def export_workbook(
    rows: list[BenchmarkRow],
    stats: SummaryStats,
    filter_description: str | None = None,
) -> bytes:
    """The filtered view as a two-sheet workbook: peer rows + summary."""
    workbook = Workbook()
    peer_sheet = workbook.active
    peer_sheet.title = "Peer comparison"
    peer_sheet.append(PEER_HEADER)
    for cell in peer_sheet[1]:
        cell.font = BOLD
        peer_sheet.column_dimensions[cell.column_letter].width = (
            PEER_COLUMN_WIDTHS.get(cell.value, DEFAULT_COLUMN_WIDTH)
        )
    for row in rows:
        peer_sheet.append(
            [
                row.name,
                row.city,
                row.state,
                row.ntee_code,
                row.total_revenue,
                row.executive_title,
                row.executive_compensation,
                row.percent_of_revenue,
                row.filing_year,
                row.data_source,
                row.propublica_url,
                row.stale,
            ]
        )
    summary_sheet = workbook.create_sheet("Summary")
    summary_sheet.column_dimensions["A"].width = DEFAULT_COLUMN_WIDTH
    summary_sheet.column_dimensions["B"].width = 2 * DEFAULT_COLUMN_WIDTH
    for label, value in [
        ("Peer count", stats.peer_count),
        ("Median compensation", stats.median),
        ("25th percentile", stats.p25),
        ("75th percentile", stats.p75),
        ("Minimum", stats.minimum),
        ("Maximum", stats.maximum),
        ("Filters", filter_description),
        ("Export date", date.today().isoformat()),
    ]:
        summary_sheet.append([label, value])
        summary_sheet.cell(row=summary_sheet.max_row, column=1).font = BOLD
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
