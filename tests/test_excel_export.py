"""Excel Export: the filtered view as a board-ready workbook (pure, no I/O)."""

from io import BytesIO

from openpyxl import load_workbook

from nonprofit_benchmark.benchmark import BenchmarkRow, SummaryStats, summarize
from nonprofit_benchmark.excel_export import export_workbook


def row(ein="111000001", **overrides):
    fields = dict(
        ein=ein,
        name=f"ORG {ein}",
        city="ALBANY",
        state="NY",
        ntee_code="A65",
        total_revenue=512000,
        executive_title="EXECUTIVE DIRECTOR",
        executive_compensation=95000,
        percent_of_revenue=18.5546875,
        filing_year=2024,
        data_source="ai",
        propublica_url=f"https://projects.propublica.org/nonprofits/organizations/{ein}",
        paid_executive_count=1,
        executives=(),
        stale=False,
    )
    fields.update(overrides)
    return BenchmarkRow(**fields)


def load(workbook_bytes):
    return load_workbook(BytesIO(workbook_bytes))


def test_export_produces_workbook_with_peer_and_summary_sheets():
    rows = [row()]

    workbook = load(export_workbook(rows, summarize(rows)))

    assert workbook.sheetnames == ["Peer comparison", "Summary"]


def test_peer_sheet_header_row_carries_every_on_screen_column_in_bold():
    rows = [row()]

    sheet = load(export_workbook(rows, summarize(rows)))["Peer comparison"]

    header = [cell.value for cell in sheet[1]]
    assert header == [
        "Organization",
        "City",
        "State",
        "NTEE",
        "Total revenue",
        "Executive title",
        "Executive compensation",
        "% of revenue",
        "Filing year",
        "Data source",
        "ProPublica URL",
        "Stale (>3 yrs)",
    ]
    assert all(cell.font.bold for cell in sheet[1])


def test_every_data_row_matches_its_input_row_with_source_and_year():
    rows = [
        row(ein="111000001", name="ARTS COUNCIL", data_source="ai"),
        row(
            ein="111000002",
            name="RIVER TRUST",
            city="TROY",
            ntee_code="C30",
            total_revenue=480000,
            executive_title=None,
            executive_compensation=120000,
            percent_of_revenue=25.0,
            filing_year=2022,
            data_source="api",
            stale=True,
        ),
    ]

    sheet = load(export_workbook(rows, summarize(rows)))["Peer comparison"]

    data = [[cell.value for cell in line] for line in sheet.iter_rows(min_row=2)]
    assert data == [
        [
            "ARTS COUNCIL", "ALBANY", "NY", "A65", 512000,
            "EXECUTIVE DIRECTOR", 95000, 18.5546875, 2024, "ai",
            "https://projects.propublica.org/nonprofits/organizations/111000001",
            False,
        ],
        [
            "RIVER TRUST", "TROY", "NY", "C30", 480000,
            None, 120000, 25.0, 2022, "api",
            "https://projects.propublica.org/nonprofits/organizations/111000002",
            True,
        ],
    ]
    # board-scrutiny requirement: every row discloses source and filing year
    assert all(line[9] in ("api", "ai") and line[8] is not None for line in data)


def test_missing_values_become_empty_cells_not_none_strings():
    failed_parse = row(
        city=None,
        ntee_code=None,
        total_revenue=None,
        executive_title=None,
        executive_compensation=None,
        percent_of_revenue=None,
    )

    sheet = load(export_workbook([failed_parse], summarize([failed_parse])))[
        "Peer comparison"
    ]

    values = [cell.value for cell in sheet[2]]
    assert values[1] is None  # city
    assert values[3] is None  # NTEE
    assert values[4] is None  # total revenue
    assert values[5] is None  # executive title
    assert values[6] is None  # executive compensation
    assert values[7] is None  # % of revenue
    assert "None" not in [v for v in values if isinstance(v, str)]


def test_summary_sheet_shows_stats_filters_and_export_date():
    from datetime import date

    rows = [
        row(ein=f"11100000{i}", executive_compensation=comp)
        for i, comp in enumerate([60000, 80000, 80000, 100000, 150000], start=1)
    ]
    stats = summarize(rows)

    sheet = load(
        export_workbook(
            rows, stats, filter_description="NY, $250K-$1M revenue, NTEE A"
        )
    )["Summary"]

    summary = {line[0].value: line[1].value for line in sheet.iter_rows()}
    assert summary["Peer count"] == 5
    assert summary["Median compensation"] == 80000
    assert summary["25th percentile"] == 80000
    assert summary["75th percentile"] == 100000
    assert summary["Minimum"] == 60000
    assert summary["Maximum"] == 150000
    assert summary["Filters"] == "NY, $250K-$1M revenue, NTEE A"
    assert summary["Export date"] == date.today().isoformat()


def test_empty_peer_set_still_produces_headers_and_summary():
    workbook = load(export_workbook([], summarize([])))

    peer_sheet = workbook["Peer comparison"]
    assert peer_sheet[1][0].value == "Organization"
    assert peer_sheet.max_row == 1  # header only, no data rows

    summary = {line[0].value: line[1].value for line in workbook["Summary"].iter_rows()}
    assert summary["Peer count"] == 0
    assert summary["Median compensation"] is None


def test_columns_are_wide_enough_to_read():
    rows = [row()]

    workbook = load(export_workbook(rows, summarize(rows)))

    peer_sheet = workbook["Peer comparison"]
    for cell in peer_sheet[1]:
        width = peer_sheet.column_dimensions[cell.column_letter].width
        assert width is not None and width >= len(str(cell.value))
    label_width = workbook["Summary"].column_dimensions["A"].width
    assert label_width is not None and label_width >= len("Median compensation")
